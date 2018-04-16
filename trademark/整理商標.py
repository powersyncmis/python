import os
import openpyxl
import pymysql.cursors
import datetime

config = {
    'host': '192.168.0.10',
    'port': 3306,
    'user': 'root',
    'password': 'power&mis',
    'db': 'mis',
    'charset': 'utf8mb4',
    'cursorclass': pymysql.cursors.DictCursor,
    }

filename = "D:\\pcpong\\商標申請專案\\群加_商標專利清單_20180409.xlsx"
if(filename):
    wb = openpyxl.load_workbook(filename, keep_vba=True, data_only=True, guess_types=True)
    for tt in ['台灣_商標清單','大陸_商標清單']:
        ws = wb.get_sheet_by_name(tt)    
        row_all=ws.max_row
        for i in range(3, row_all+1):
            if ws.cell(row=i, column=1).value is None:
                break
            var1 = ws.cell(row=i, column=1).value # 類別
            var2 = ws.cell(row=i, column=4).value # 証書編號
            var3 = ws.cell(row=i, column=2).value # 名稱
            if ws.cell(row=i, column=5).value is None:
                var4 = ''
            else:
                var4 = datetime.datetime.strftime(ws.cell(row=i, column=5).value, '%Y/%m/%d') # 起始日期
            if ws.cell(row=i, column=6).value is None:
                var5 = ''
            else:
                var5 = datetime.datetime.strftime(ws.cell(row=i, column=6).value, '%Y/%m/%d') # 到期日期
            var6 = ws.cell(row=i, column=7).value # 所有權人
            var7 = ws.cell(row=i, column=8).value # 備註
            print(i)
            connection4 = pymysql.connect( ** config)
            try:
                with connection4.cursor() as cursor4:
                    sql4 = "INSERT trademark(Type,Class,SNo,SName,Start_Date,End_Date,Owner,Location,Other,Who) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
                    # sql4 = "INSERT INTO trademark(Type,Class) VALUES(%s,%s)"
                    cursor4.execute(sql4,('0',var1,var2,var3,var4,var5,var6,tt[0:2],var7,'彭彥碩'))
                    connection4.commit()
            except:
                print("level_3")
            finally:
                cursor4.close()
                connection4.close()

    for tt in ['美_日商標清單']:
        ws = wb.get_sheet_by_name(tt)    
        row_all=ws.max_row
        for i in range(3, row_all+1):
            if ws.cell(row=i, column=1).value is None:
                break
            var0 = ws.cell(row=i, column=1).value # 地區
            var1 = ws.cell(row=i, column=2).value # 類別
            var2 = ws.cell(row=i, column=5).value # 証書編號
            var3 = ws.cell(row=i, column=3).value # 名稱
            if ws.cell(row=i, column=6).value is None:
                var4 = ''
            else:
                var4 = datetime.datetime.strftime(ws.cell(row=i, column=6).value, '%Y/%m/%d') # 起始日期
            if ws.cell(row=i, column=7).value is None:
                var5 = ''
            else:
                var5 = datetime.datetime.strftime(ws.cell(row=i, column=7).value, '%Y/%m/%d') # 到期日期
            var6 = ws.cell(row=i, column=8).value # 所有權人
            var7 = ws.cell(row=i, column=9).value # 備註
            print(i)
            connection4 = pymysql.connect( ** config)
            try:
                with connection4.cursor() as cursor4:
                    sql4 = "INSERT trademark(Type,Class,SNo,SName,Start_Date,End_Date,Owner,Location,Other,Who) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
                    # sql4 = "INSERT INTO trademark(Type,Class) VALUES(%s,%s)"
                    cursor4.execute(sql4,('0',var1,var2,var3,var4,var5,var6,var0,var7,'彭彥碩'))
                    connection4.commit()
            except:
                print("level_3")
            finally:
                cursor4.close()
                connection4.close()
    
    for tt in ['專利清單']:
        ws = wb.get_sheet_by_name(tt)    
        row_all=ws.max_row
        for i in range(3, row_all+1):
            if ws.cell(row=i, column=1).value is None:
                break
            var0 = ws.cell(row=i, column=1).value # 地區
            var2 = ws.cell(row=i, column=3).value # 証書編號
            var3 = ws.cell(row=i, column=2).value # 名稱
            if ws.cell(row=i, column=5).value is None:
                var4 = ''
            else:
                var4 = datetime.datetime.strftime(ws.cell(row=i, column=5).value, '%Y/%m/%d') # 起始日期
            if ws.cell(row=i, column=6).value is None:
                var5 = ''
            else:
                var5 = datetime.datetime.strftime(ws.cell(row=i, column=6).value, '%Y/%m/%d') # 到期日期
            var6 = ws.cell(row=i, column=7).value # 所有權人
            var7 = ws.cell(row=i, column=10).value # 備註
            if ws.cell(row=i, column=8).value is None:
                var8 = ''
            else:
                var8 = datetime.datetime.strftime(ws.cell(row=i, column=8).value, '%Y/%m/%d') # 應繳日期
            var9 = ws.cell(row=i, column=9).value # 金額
            print(i)
            connection4 = pymysql.connect( ** config)
            try:
                with connection4.cursor() as cursor4:
                    sql4 = "INSERT trademark(Type,SNo,SName,Start_Date,End_Date,Owner,Pay_Date,Cash,Location,Other,Who) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
                    # sql4 = "INSERT INTO trademark(Type,Class) VALUES(%s,%s)"
                    cursor4.execute(sql4,('1',var2,var3,var4,var5,var6,var8,var9,var0,var7,'彭彥碩'))
                    connection4.commit()
            except Exception as e:
                print(e)
            finally:
                cursor4.close()
                connection4.close()