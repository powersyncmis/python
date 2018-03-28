import sys
import openpyxl
import tkinter as tk
from tkinter import Tk, filedialog
import pyexcel as p
from datetime import datetime
import pyodbc

x=datetime.now().strftime("%Y%m%d")
non_bmp_map = dict.fromkeys(range(0x10000, sys.maxunicode + 1), 0xfffd)
root = tk.Tk()
root.withdraw()
filename = filedialog.askopenfilename()
if filename[-4:]==".xls":
    newfile=filename[0:-4]+'.xlsx'
    p.save_book_as(file_name=filename,dest_file_name=newfile)
    filename=newfile

if(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.get_sheet_by_name('Sheet1')
    row_all = ws.rows
    wb1 = openpyxl.Workbook()
    ws1 = wb1.get_active_sheet()

    ws1.cell(row=1,column=1).value = '客戶'
    ws1.cell(row=1,column=2).value = '訂單號'
    ws1.cell(row=1,column=3).value = '品號'
    ws1.cell(row=1,column=4).value = '數量'
    ws1.cell(row=1,column=5).value = '尾數'
    ws1.cell(row=1,column=6).value = '庫別'
    ws1.cell(row=1,column=7).value = '主諸位'
    ws1.cell(row=1,column=8).value = '位置'
    ws1.cell(row=1,column=9).value = '庫存數'
    
    j = 1
    i = 1
    z = 2
    zz = 0
    for rows in row_all:
        if ws.cell(row=j,column=1).value=="請 購／訂 購 單":
            m01 = str(ws.cell(row=j,column=9).value).translate(non_bmp_map)
            m02 = str(ws.cell(row=j,column=20).value).translate(non_bmp_map)
        if (ws.cell(row=j,column=1).value==1 or ws.cell(row=j,column=1).value==2 or ws.cell(row=j,column=1).value==3 or 
            ws.cell(row=j,column=1).value==4 or ws.cell(row=j,column=1).value==5 or ws.cell(row=j,column=1).value==6 or 
            ws.cell(row=j,column=1).value==7 or ws.cell(row=j,column=1).value==8 or ws.cell(row=j,column=1).value==9 or 
            ws.cell(row=j,column=1).value==10):
            ws1.cell(row=z,column=1).value = m01
            ws1.cell(row=z,column=2).value = m02
            ws1.cell(row=z,column=3).value = str(ws.cell(row=j,column=2).value).translate(non_bmp_map)
            if str(ws.cell(row=j,column=14).value).translate(non_bmp_map)=='NULL':
                ws1.cell(row=z,column=4).value = str(ws.cell(row=j,column=19).value).translate(non_bmp_map)
            else:
                ws1.cell(row=z,column=4).value = str(ws.cell(row=j,column=20).value).translate(non_bmp_map)
            
            z=z+1
            zz=zz+1
        j=j+1
    ws1.cell(row=z-1,column=5).value =  str(ws1.cell(row=z-1,column=2).value)[-3:]
    filename1="c:\TEMP\寶雅訂單_{0}.xlsx".format(x)
    wb1.save(filename=filename1)
    print('整理完成！總計{0}筆商品。'.format(zz))

wb = openpyxl.load_workbook(filename1)
ws = wb.get_sheet_by_name('Sheet')
row_all = ws.rows


db = pyodbc.connect('DRIVER={SQL Server};SERVER=192.168.0.5;DATABASE=Leader;UID=sa;PWD=dsc@54302158')
cursor1 = db.cursor()
cursor2 = db.cursor()
z=2
for rows in row_all:
    try:
        select_table_sql1="""SELECT INVMC.MC001, INVMC.MC002, INVMC.MC015, INVMM.MM003, INVMM.MM005 
                            FROM INVMC INNER JOIN INVMM ON INVMM.MM001 = INVMC.MC001 
                            WHERE (INVMM.MM002 = 'A01' AND INVMC.MC002='A01') AND (INVMC.MC001 = '%s' AND INVMM.MM005 > 0) ORDER BY INVMM.MM005""" % (str(ws.cell(row=z,column=3).value).translate(non_bmp_map))
        cursor1.execute(select_table_sql1)
        results1 = cursor1.fetchall()
        if results1 is None:
            print("1")
            z = z + 1
        else:
            for i in results1:
                ws.cell(row=z,column=6).value=i[1]
                ws.cell(row=z,column=7).value=i[2]
                ws.cell(row=z,column=8).value=i[3]
                ws.cell(row=z,column=9).value=i[4]
                ws.cell(row=z,column=10).value="1"
    finally:
        z = z + 1
db.close()
wb.save(filename1)
print("處理完畢")