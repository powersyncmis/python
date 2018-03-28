import sys
import openpyxl
import tkinter as tk
from tkinter import Tk, filedialog
import pyexcel as p

non_bmp_map = dict.fromkeys(range(0x10000, sys.maxunicode + 1), 0xfffd)
root = tk.Tk()
root.withdraw()
filename = filedialog.askopenfilename()

if filename[-4:]==".xls":
    newfile=filename[0:-4]+'.xlsx'
    p.save_book_as(file_name=filename,dest_file_name=newfile)
    filename=newfile

if(filename):
    wb = openpyxl.load_workbook(filename, keep_vba=True)
    ws = wb.get_sheet_by_name('orders')
    row_all = ws.rows
    wb1 = openpyxl.Workbook()
    ws1 = wb1.get_active_sheet()

    ws1.cell(row=1,column=1).value = '客戶單號'
    ws1.cell(row=1,column=2).value = '買家帳號'
    ws1.cell(row=1,column=3).value = '買家支付的運費'
    ws1.cell(row=1,column=4).value = '商品品號'
    ws1.cell(row=1,column=5).value = '訂單數量'
    ws1.cell(row=1,column=6).value = '收件地址'
    ws1.cell(row=1,column=7).value = '收件者姓名'
    ws1.cell(row=1,column=8).value = '電話號碼'
    ws1.cell(row=1,column=9).value = '寄送方式'
    ws1.cell(row=1,column=10).value = '包裹查詢號碼'
    ws1.cell(row=1,column=11).value = '買家備註'
    
    i = 0
    j = 1
    jj = 1
    xx = []
    for rows in row_all:
        if str(ws.cell(row=j,column=1).value).strip()=="訂單編號":
            j=j+1
            jj=jj+1
            pass
        elif ws.cell(row=j,column=1).value!=None:
            for x in range(1,29):
                if x == 10:
                    times=str(ws.cell(row=j,column=x).value).count("[")
                    z=0
                    z=len(str(ws.cell(row=j,column=x).value))
                    y=0
                    x1=0
                    x2=0
                    x4=0
                    x5=0
                    while x5<z:
                        y=y+x2
                        x1=str(ws.cell(row=j,column=x).value).find(":",y)+1
                        x2=str(ws.cell(row=j,column=x).value).find(";",x1)
                        x5=str(ws.cell(row=j,column=x).value).find(";",x2)+3
                        x3=str(ws.cell(row=j,column=x).value)[x1:x2].translate(non_bmp_map)
                        i=i+1
                        xx.append(x3.strip())
                        y=0
                    #print(xx,len(xx),times)
                    for f in range(0,len(xx)):
                        if "$ " in xx[f] and "價$" not in xx[f]:
                            ws1.cell(row=jj,column=1).value=str(ws.cell(row=j,column=1).value).translate(non_bmp_map)
                            ws1.cell(row=jj,column=2).value=str(ws.cell(row=j,column=4).value).translate(non_bmp_map)
                            ws1.cell(row=jj,column=3).value=float(ws.cell(row=j,column=8).value)
                            ws1.cell(row=jj,column=4).value=xx[f+2]
                            ws1.cell(row=jj,column=5).value=xx[f+1]
                            ws1.cell(row=jj,column=6).value=str(ws.cell(row=j,column=11).value).translate(non_bmp_map)
                            ws1.cell(row=jj,column=7).value=str(ws.cell(row=j,column=16).value).translate(non_bmp_map)
                            ws1.cell(row=jj,column=8).value=str(ws.cell(row=j,column=17).value).translate(non_bmp_map)
                            ws1.cell(row=jj,column=9).value=str(ws.cell(row=j,column=18).value).translate(non_bmp_map)
                            if ws.cell(row=j,column=24).value==None:
                                ws1.cell(row=jj,column=10).value=" "
                            else:
                                ws1.cell(row=jj,column=10).value=str(ws.cell(row=j,column=24).value).translate(non_bmp_map)
                                ws1.cell(row=jj,column=11).value=str(ws.cell(row=j,column=28).value).translate(non_bmp_map)
                            jj=jj+1
                    j=j+1
                    xx = []
        
            #z=z+1
        #else:
         #   for y in range(8,83):
          #      
           #     if ws.cell(row=j,column=y).value==None and ws.cell(row=j,column=y+1).value==None:
            #        continue
             #   else:
              #      zz=(y-2)%5
               #     
                #    if(zz==0):
                 #       ws1.cell(row=z, column=5).value = str(ws.cell(row=j,column=y).value).translate(non_bmp_map)
                  #      z=z+1
                   # else:
                    #    ws1.cell(row=z, column=zz).value = str(ws.cell(row=j,column=y).value).translate(non_bmp_map)
                        
    wb1.save(filename=filename[-13:])
    print('整理完成！總計{0}筆訂單。'.format(jj-1))
