import sys
import openpyxl
import tkinter as tk
from tkinter import Tk, filedialog


non_bmp_map = dict.fromkeys(range(0x10000, sys.maxunicode + 1), 0xfffd)
root = tk.Tk()
root.withdraw()
filename = filedialog.askopenfilename()

if(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.get_sheet_by_name('Sheet1')
    row_all = ws.rows
    wb1 = openpyxl.Workbook()
    ws1 = wb1.get_active_sheet()

    ws1.cell(row=1,column=1).value = '商品ID'
    ws1.cell(row=1,column=2).value = '商品貨號'
    ws1.cell(row=1,column=3).value = '商品名稱'
    ws1.cell(row=1,column=4).value = '價格'
    ws1.cell(row=1,column=5).value = '庫存'
    
    j = 3
    i = 1
    z = 2
    zz =1
    for rows in row_all:
        j=j+1
        if ws.cell(row=j,column=7).value!=None:
            i = 1
            for x in range(1,8):
                if x == 4 or x == 5:
                    continue
                else:
                    ws1.cell(row=z,column=i).value = str(ws.cell(row=j,column=x).value).translate(non_bmp_map)
                    i=i+1
            z=z+1
        else:
            for y in range(8,83):
                
                if ws.cell(row=j,column=y).value==None and ws.cell(row=j,column=y+1).value==None:
                    continue
                else:
                    zz=(y-2)%5
                    
                    if(zz==0):
                        ws1.cell(row=z, column=5).value = str(ws.cell(row=j,column=y).value).translate(non_bmp_map)
                        z=z+1
                    else:
                        ws1.cell(row=z, column=zz).value = str(ws.cell(row=j,column=y).value).translate(non_bmp_map)
                        
    wb1.save(filename=filename[-15:])
    print('整理完成！總計{0}筆商品。'.format(z))
