import os
import openpyxl
import pymysql.cursors
import datetime

config = {
    'host': '192.168.0.10',
    'port': 3306,
    'user': 'root',
    'password': 'power&mis',
    'db': 'mis',
    'charset': 'utf8mb4',
    'cursorclass': pymysql.cursors.DictCursor,
    }

filename = "D:\\pcpong\\國際條碼\\條碼確認檔-網路新增-201711215.xlsx"
if(filename):
    wb = openpyxl.load_workbook(filename, keep_vba=True, data_only=True)
    ws = wb.get_sheet_by_name('2018台北')    
    row_all=ws.max_row
    for i in range(6, row_all):
        if ws.cell(row=i, column=1).value is None:
            break
        # var1 = ws.cell(row=i, column=4).value # 條碼
        # var2 = ws.cell(row=i, column=6).value # 規格
        # var3 = ws.cell(row=i, column=7).value # 包裝方式
        var1 = datetime.datetime.strftime(ws.cell(row=i, column=1).value, '%Y/%m/%d') # 日期
        var2 = ws.cell(row=i, column=2).value # 品號
        var3 = ws.cell(row=i, column=4).value # 條碼
        var4 = ws.cell(row=i, column=6).value # 規格
        var5 = ws.cell(row=i, column=7).value # 包裝
        print(i)
        # print(var1,var2,var3)
        connection4 = pymysql.connect( ** config)
        try:
            with connection4.cursor() as cursor4:
                # sql4 = "UPDATE gs1 SET SName = %s, Spec = %s, Company = %s WHERE GS1 = %s"
                sql4 = "INSERT gs1(Create_Date,SNo,GS1,SName,Spec,End_Case,Company) Values(%s,%s,%s,%s,%s,%s,%s)"
                cursor4.execute(sql4,(var1,var2,var3,var4,var5,'1','TP'))
                connection4.commit()
        except:
            print("level_3")
        finally:
            cursor4.close()
            connection4.close()